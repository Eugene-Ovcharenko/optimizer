from openpyxl import load_workbook
import pandas as pd
import numpy as np
from glob2 import glob
import datetime
import random
import open3d as o3d
import trimesh
import pathlib
from random import random
from utils.global_variable import (get_id, set_id, get_problem_name, get_mesh_step,
                                   get_cpus, get_base_name, get_s_lim, get_direction,
                                   change_direction, get_valve_position)
from pymoo.problems import get_problem
from utils.get_history_output import get_history_output as get_history_output
from utils.run_abaqus import run_abaqus as run_abaqus
from utils.read_data import read_data
from utils.purgeFiles import purgeFiles
from utils.logger_leaflet import configure_log_leaflet, cleanup_log_leaflet
from utils.generateShell import generateShell
from utils.createGeometry import createGeometry
from utils.logger_leaflet import log_message
from utils.write_inp import write_inp_shell, write_inp_contact
import os

now = str(datetime.datetime.now()).replace(' ', '_').replace(':', '-').split('.')[0]
now = now[:-3]

pathToAbaqus = str(pathlib.Path(__file__).parent.resolve()) + '/abaqusWF/'
path_utils = str(pathlib.Path(__file__).parent.resolve())

# procedure parameters class
class Procedure:
    baseName = None
    cpus = 4
    logging = True
    wbResults = None
    wbLog = None
    outFileNameResults = None
    outFileNameLog = None
    sheet_short = 'short'
    mesh_step = None

    def __init__(self, cpus, logging, baseName, wbResults, wbLog, outFileNameResults, outFileNameLog,
                 sheet_short, sheet_desc, mesh_step):
        self.baseName = baseName
        self.cpus = cpus
        self.logging = logging
        self.wbResults = wbResults
        self.wbLog = wbLog
        self.outFileNameResults = outFileNameResults
        self.outFileNameLog = outFileNameLog
        self.sheet_short = sheet_short
        self.sheet_desc = sheet_desc
        self.mesh_step = mesh_step

    # ==================================================================================================================
    # __________________________________________RUN PROCEDURE___________________________________________________________
    # ==================================================================================================================

    def run_procedure(self, params) -> dict:

        def run_leaflet_single(self, params) -> dict:
            ID = get_id()
            try:
                # log_message(f"current ID is {get_id()}. current time is {str(datetime.datetime.now()).split('.')[0]}\n")
                baseName = self.baseName + '_' + now + '_' + str(ID)
                tt1 = datetime.datetime.now()
                try:
                    HGT, Lstr, THK, ANG, CVT, LAS = params
                except:
                    Lstr, ANG, CVT, LAS = params
                    HGT = 11
                    THK = 0.3
                DIA = 22.98
                Lift = 0
                SEC = 120
                EM = 3.2
                mesh_step = self.mesh_step
                tangent_behavior = 0.05
                normal_behavior = 0.2
                fileName = baseName + '.inp'
                try:
                    pointsInner, _, _, _, pointsHullLower, _, points, _, finalRad, currRad, message = \
                        createGeometry(HGT=HGT, Lstr=Lstr, SEC=SEC, DIA=DIA, THK=0.5,
                                       ANG=ANG, Lift=Lift, CVT=CVT, LAS=LAS, mesh_step=mesh_step)
                except Exception as e:
                    delta = datetime.datetime.now() - tt1
                    # wbGeom = load_workbook(filename=self.outFileNameResults)
                    sheet = self.wbLog['log']
                    sheet.append(
                        [fileName, ID, HGT, Lstr, SEC, DIA, THK, ANG, Lift, CVT, LAS, EM, tangent_behavior,
                         normal_behavior,
                         0, str(e), delta])
                    self.wbLog.save(self.outFileNameLog)
                    self.wbLog.close()
                    del delta, sheet
                    raise e

                k = 1.1
                flag_calk_k = True
                while flag_calk_k:
                    try:
                        o3d.utility.set_verbosity_level(o3d.utility.VerbosityLevel(0))
                        pcd = o3d.geometry.PointCloud()
                        pcd.points = o3d.utility.Vector3dVector(points.T)
                        mesh = o3d.geometry.TriangleMesh.create_from_point_cloud_alpha_shape(pcd, alpha=0.5 * k)
                        _ = o3d.io.write_triangle_mesh('./utils/geoms/temp_' + '.ply', mesh, write_vertex_normals=True)

                        mesh = trimesh.load_mesh('./utils/geoms/temp_' + '.ply')
                        mesh.fix_normals()  # fix wrong normals
                        mesh.export('./utils/geoms/temp_' + '.stl')

                        flag_calk_k = False
                    except:
                        k += 0.1

                tt2 = datetime.datetime.now()

                try:
                    shellNode, shellEle, fixed_bc = generateShell(points=mesh.vertices, elements=mesh.faces,
                                                                  pointsInner=pointsInner,
                                                                  pointsHullLower=pointsHullLower, meshStep=mesh_step)
                    tt2 = datetime.datetime.now()
                    message = 'done'
                    log_message(
                        message + '. Mesh nodes is ' + str(len(shellNode)) + '. Elements is ' + str(len(shellEle)))

                except Exception as e:
                    delta = datetime.datetime.now() - tt1
                    sheet = self.wbLog['log']
                    sheet.append(
                        [fileName, ID, HGT, Lstr, SEC, DIA, THK, ANG, Lift, CVT, LAS, EM, tangent_behavior,
                         normal_behavior, 0, str(e), delta])
                    self.wbLog.save(self.outFileNameLog)
                    self.wbLog.close()
                    del delta, sheet, pcd
                    raise e

                del mesh
                inpFileName = str(inpDir) + str(baseName)
                jobName = str(baseName) + '_Job'
                modelName = str(baseName) + '_Model'
                partName = str(baseName) + '_Part'
                outFEATime = 0
                if get_direction().lower() == 'direct':
                    try:
                        write_inp_shell(
                            fileName=inpFileName + '.inp', Nodes=shellNode, Elements=shellEle,
                            BCfix=fixed_bc, THC=THK, Emod=EM, Dens=9e-10, JobName=jobName,
                            ModelName=modelName, partName=partName, MaterialName='PVA', PressType='vent',
                            press_overclosure='linear', tangent_behavior=tangent_behavior,
                            normal_behavior=normal_behavior
                        )
                        message = run_abaqus(pathToAbaqus, jobName, inpFileName, self.cpus)
                        outFEATime = datetime.datetime.now() - tt1
                    except Exception as e:
                        delta = datetime.datetime.now() - tt1
                        # wbGeom = load_workbook(filename=self.outFileNameLog)
                        sheet = self.wbLog['log']
                        sheet.append(
                            [fileName, ID, HGT, Lstr, SEC, DIA, THK, ANG, Lift, CVT, LAS, EM, tangent_behavior,
                             normal_behavior,
                             0, str(e), delta])
                        self.wbLog.save(self.outFileNameLog)
                        self.wbLog.close()
                        del delta, sheet
                        raise e

                    # парсим odb, считываем поля, считаем максимумы и площадь открытия, пишем в outFileName
                    try:
                        get_history_output(pathName=pathToAbaqus, odbFileName=jobName + '.odb', cpus=self.cpus)
                    except:
                        delta = datetime.datetime.now() - tt1
                        # wbGeom = load_workbook(filename=self.outFileNameLog)
                        sheet = self.wbLog['log']
                        with open(pathToAbaqus + '/results/' + partName[0:-5] + '/FramesCount.txt', 'r') as DataFile:
                            frames = int(DataFile.read())
                        sheet.append(
                            [fileName, ID, HGT, Lstr, SEC, DIA, THK, ANG, Lift, CVT, LAS, EM, tangent_behavior,
                             normal_behavior,
                             frames,
                             'Odb parse problem', delta])
                        self.wbLog.save(self.outFileNameLog)
                        self.wbLog.close()
                        # purgeFiles(pathToAbaqus + 'results/', partName, pathToAbaqus, jobName)
                        del delta, sheet
                        raise 'Odb parse problem'

                    try:
                        endPath = pathToAbaqus + 'results/'
                        LMN_op, LMN_cl, Smax, VMS, perf_index = read_data(
                            pathToAbaqus=pathToAbaqus, endPath=endPath, partName=partName,
                            HGT=HGT, Lstr=Lstr, DIA=DIA, THK=THK, ANG=ANG, Lift=Lift, CVT=CVT, LAS=LAS, EM=EM, SEC=SEC,
                            tangent_behavior=tangent_behavior, normal_behavior=normal_behavior,
                            mesh_step=mesh_step, outFEATime=outFEATime, fileName=fileName,
                            sheet_short=self.sheet_short, sheet_desc=self.sheet_desc, wbResults=self.wbResults,
                            outFileNameResult=self.outFileNameResults, outFileNameGeom=self.outFileNameLog, tt1=tt1
                        )
                        purgeFiles(endPath, partName, pathToAbaqus, jobName)
                    except Exception as e:
                        delta = datetime.datetime.now() - tt1
                        sheet = self.wbLog['log']
                        try:
                            with open(pathToAbaqus + '/results/' + partName[0:-5].upper() + '/FramesCount.txt',
                                      'r') as DataFile:
                                frames = int(DataFile.read())
                        except:
                            frames = 0
                        sheet.append(
                            [fileName, ID, HGT, Lstr, SEC, DIA, THK, ANG, Lift, CVT, LAS, EM, tangent_behavior,
                             normal_behavior,
                             frames,
                             str(e), delta])
                        self.wbLog.save(self.outFileNameLog)
                        self.wbLog.close()
                        purgeFiles(pathToAbaqus + 'results/', partName, pathToAbaqus, jobName)
                        del delta, sheet
                        raise e

                    if LMN_op < 0:
                        change_direction()
                        log_message(f'Changed direction to {get_direction()}! LMN_op is {LMN_op}\m')
                        try:
                            write_inp_shell(
                                fileName=inpFileName + '.inp', Nodes=shellNode, Elements=shellEle,
                                BCfix=fixed_bc, THC=THK, Emod=EM, Dens=9e-10, JobName=jobName,
                                ModelName=modelName, partName=partName, MaterialName='PVA', PressType='vent',
                                press_overclosure='linear', tangent_behavior=tangent_behavior,
                                normal_behavior=normal_behavior
                            )
                            message = run_abaqus(pathToAbaqus, jobName, inpFileName, self.cpus)
                            outFEATime = datetime.datetime.now() - tt1
                        except Exception as e:
                            delta = datetime.datetime.now() - tt1
                            # wbGeom = load_workbook(filename=self.outFileNameLog)
                            sheet = self.wbLog['log']
                            sheet.append(
                                [fileName, ID, HGT, Lstr, SEC, DIA, THK, ANG, Lift, CVT, LAS, EM, tangent_behavior,
                                 normal_behavior,
                                 0, str(e), delta])
                            self.wbLog.save(self.outFileNameLog)
                            self.wbLog.close()
                            del delta, sheet
                            raise e

                        # парсим odb, считываем поля, считаем максимумы и площадь открытия, пишем в outFileName
                        try:
                            get_history_output(pathName=pathToAbaqus, odbFileName=jobName + '.odb', cpus=self.cpus)
                        except:
                            delta = datetime.datetime.now() - tt1
                            # wbGeom = load_workbook(filename=self.outFileNameLog)
                            sheet = self.wbLog['log']
                            with open(pathToAbaqus + '/results/' + partName[0:-5] + '/FramesCount.txt', 'r') as DataFile:
                                frames = int(DataFile.read())
                            sheet.append(
                                [fileName, ID, HGT, Lstr, SEC, DIA, THK, ANG, Lift, CVT, LAS, EM, tangent_behavior,
                                 normal_behavior,
                                 frames,
                                 'Odb parse problem', delta])
                            self.wbLog.save(self.outFileNameLog)
                            self.wbLog.close()
                            # purgeFiles(pathToAbaqus + 'results/', partName, pathToAbaqus, jobName)
                            del delta, sheet
                            raise 'Odb parse problem'

                        try:
                            endPath = pathToAbaqus + 'results/'
                            LMN_op, LMN_cl, Smax, VMS, perf_index = read_data(
                                pathToAbaqus=pathToAbaqus, endPath=endPath, partName=partName,
                                HGT=HGT, Lstr=Lstr, DIA=DIA, THK=THK, ANG=ANG, Lift=Lift, CVT=CVT, LAS=LAS, EM=EM, SEC=SEC,
                                tangent_behavior=tangent_behavior, normal_behavior=normal_behavior,
                                mesh_step=mesh_step, outFEATime=outFEATime, fileName=fileName,
                                sheet_short=self.sheet_short, sheet_desc=self.sheet_desc, wbResults=self.wbResults,
                                outFileNameResult=self.outFileNameResults, outFileNameGeom=self.outFileNameLog, tt1=tt1
                            )
                            purgeFiles(endPath, partName, pathToAbaqus, jobName)
                        except Exception as e:
                            delta = datetime.datetime.now() - tt1
                            sheet = self.wbLog['log']
                            try:
                                with open(pathToAbaqus + '/results/' + partName[0:-5].upper() + '/FramesCount.txt',
                                          'r') as DataFile:
                                    frames = int(DataFile.read())
                            except:
                                frames = 0
                            sheet.append(
                                [fileName, ID, HGT, Lstr, SEC, DIA, THK, ANG, Lift, CVT, LAS, EM, tangent_behavior,
                                 normal_behavior,
                                 frames,
                                 str(e), delta])
                            self.wbLog.save(self.outFileNameLog)
                            self.wbLog.close()
                            purgeFiles(pathToAbaqus + 'results/', partName, pathToAbaqus, jobName)
                            del delta, sheet
                            raise e

                        change_direction()
                        log_message(f'Swap back direction to {get_direction()}! LMN_op is {LMN_op}')

                del fixed_bc, partName, jobName, endPath, modelName, inpFileName
                del tt1, tt2

                if LMN_cl < 0:
                    out_lmn_cl = 0
                else:
                    out_lmn_cl = LMN_cl

                objectives_dict = {
                    '1 - LMN_open': 1 - LMN_op,
                    'LMN_open': LMN_op,
                    "LMN_closed": out_lmn_cl,
                    "Smax": Smax
                }

                constraints_dict = {
                    "VMS-Smax": VMS - get_s_lim()
                }

                # cleanup_log_leaflet()
                # set_id(ID + 1)

                return {"objectives": objectives_dict, "constraints": constraints_dict}
            except Exception as exept:
                log_message(f'Exception: {exept}')

                objectives_dict = {
                    'LMN_open': 0.0,
                    '1 - LMN_open': 1.0,
                    "LMN_closed": 1.0,
                    "Smax": 5
                }

                constraints_dict = {
                    "VMS-Smax": 5
                }

                set_dead_objects(get_dead_objects() + 1)
                return {"objectives": objectives_dict, "constraints": constraints_dict}

        def run_leaflet_contact(self, params) -> dict:
            log_message(f"current ID is {get_id()}. current time is {str(datetime.datetime.now()).split('.')[0]}\n")
            try:
                log_message(f"current ID is {get_id()}\n")
                baseName = self.baseName + '_' + now
                tt1 = datetime.datetime.now()
                try:
                    HGT, Lstr, THK, ANG, CVT, LAS = params
                except:
                    Lstr, ANG, CVT, LAS = params
                    HGT = 11
                    THK = 0.3
                DIA = 22.98
                Lift = 0
                EM = 1.88 # Formlabs elastic 50A
                mesh_step = self.mesh_step
                tangent_behavior = 1
                normal_behavior = 0.2
                SEC = 119
                Dens = 1.02e-9
                MaterialName = 'FormLabs Elasctic 50A'
                PressType = get_valve_position()  # can be 'vent'
                fileName = baseName + '.inp'
                try:
                    pointsInner, _, _, _, pointsHullLower, _, points, _, finalRad, currRad, message = \
                        createGeometry(HGT=HGT, Lstr=Lstr, SEC=SEC, DIA=DIA, THK=0.35,
                                       ANG=ANG, Lift=Lift, CVT=CVT, LAS=LAS, mesh_step=mesh_step)
                except Exception as e:
                    delta = datetime.datetime.now() - tt1
                    sheet = self.wbLog['log']
                    sheet.append(
                        [fileName, get_id(), HGT, Lstr, SEC, DIA, THK, ANG, Lift, CVT, LAS, EM, tangent_behavior,
                         normal_behavior,
                         0, str(e), delta])
                    self.wbLog.save(self.outFileNameLog)
                    self.wbLog.close()
                    del delta, sheet
                    raise e

                k = 1.1
                flag_calk_k = True
                while flag_calk_k:
                    try:
                        o3d.utility.set_verbosity_level(o3d.utility.VerbosityLevel(0))
                        pcd = o3d.geometry.PointCloud()
                        pcd.points = o3d.utility.Vector3dVector(points.T)
                        mesh = o3d.geometry.TriangleMesh.create_from_point_cloud_alpha_shape(pcd, alpha=0.5 * k)
                        _ = o3d.io.write_triangle_mesh(path_utils+'/geoms/temp_' + '.ply', mesh, write_vertex_normals=True)

                        mesh = trimesh.load_mesh(path_utils + '/geoms/temp_' + '.ply')
                        mesh.fix_normals()  # fix wrong normals
                        mesh.export(path_utils + '/geoms/temp_' + '.stl')

                        flag_calk_k = False
                    except:
                        k += 0.1

                tt2 = datetime.datetime.now()

                try:
                    shellNode, shellEle, fixed_bc = generateShell(points=mesh.vertices, elements=mesh.faces,
                                                                  pointsInner=pointsInner,
                                                                  pointsHullLower=pointsHullLower, meshStep=mesh_step)
                    tt2 = datetime.datetime.now()
                    message = 'done'
                    log_message(
                        message + '. Mesh nodes is ' + str(len(shellNode)) + '. Elements is ' + str(len(shellEle)))

                except Exception as e:
                    delta = datetime.datetime.now() - tt1
                    sheet = self.wbLog['log']
                    sheet.append(
                        [fileName, get_id(), HGT, Lstr, SEC, DIA, THK, ANG, Lift, CVT, LAS, EM, tangent_behavior,
                         normal_behavior,
                         0, str(e), delta])
                    self.wbLog.save(self.outFileNameLog)
                    self.wbLog.close()
                    del delta, sheet, pcd
                    raise e

                del mesh
                inpFileName = str(inpDir) + str(baseName)
                jobName = str(baseName) + '_Job'
                modelName = str(baseName) + '_Model'
                partName = str(baseName) + '_Part'
                outFEATime = 0

                try:
                    write_inp_contact(
                        fileName=inpFileName + '.inp', Nodes=shellNode, Elements=shellEle,
                        BCfix=fixed_bc, THC=THK, Emod=EM, Dens=Dens, JobName=jobName,
                        ModelName=modelName, partName=partName, MaterialName=MaterialName, PressType=PressType,
                        press_overclosure='linear', tangent_behavior=tangent_behavior,
                        normal_behavior=normal_behavior
                    )
                    message = run_abaqus(pathToAbaqus, jobName, inpFileName, self.cpus)
                    outFEATime = datetime.datetime.now() - tt1
                except Exception as e:
                    delta = datetime.datetime.now() - tt1
                    # wbGeom = load_workbook(filename=self.outFileNameLog)
                    sheet = self.wbLog['log']
                    sheet.append(
                        [fileName, get_id(), HGT, Lstr, SEC, DIA, THK, ANG, Lift, CVT, LAS, EM, tangent_behavior,
                         normal_behavior,
                         0, str(e), delta])
                    self.wbLog.save(self.outFileNameLog)
                    self.wbLog.close()
                    del delta, sheet
                    raise e

                # парсим odb, считываем поля, считаем максимумы и площадь открытия, пишем в outFileName
                try:
                    _ = get_history_output(pathName=pathToAbaqus, odbFileName=jobName + '.odb')
                except:
                    delta = datetime.datetime.now() - tt1
                    # wbGeom = load_workbook(filename=self.outFileNameLog)
                    sheet = self.wbLog['log']
                    with open(pathToAbaqus + '/results/' + partName[0:-5] + '/FramesCount.txt', 'r') as DataFile:
                        frames = int(DataFile.read())
                    sheet.append(
                        [fileName, get_id(), HGT, Lstr, SEC, DIA, THK, ANG, Lift, CVT, LAS, EM, tangent_behavior,
                         normal_behavior,
                         frames,
                         'Odb parse problem', delta])
                    self.wbLog.save(self.outFileNameLog)
                    self.wbLog.close()
                    # purgeFiles(pathToAbaqus + 'results/', partName, pathToAbaqus, jobName)
                    del delta, sheet
                    raise 'Odb parse problem'

                try:
                    endPath = pathToAbaqus + 'results/'
                    LMN_op, LMN_cl, Smax, VMS, perf_index, heli = read_data(
                        pathToAbaqus=pathToAbaqus, endPath=endPath, partName=partName, Slim=3.23,
                        HGT=HGT, Lstr=Lstr, DIA=DIA, THK=THK, ANG=ANG, Lift=Lift, CVT=CVT, LAS=LAS, EM=EM, SEC=SEC,
                        tangent_behavior=tangent_behavior, normal_behavior=normal_behavior,
                        mesh_step=mesh_step, outFEATime=outFEATime, fileName=fileName,
                        sheet_short=self.sheet_short, sheet_desc=self.sheet_desc, wbResults=self.wbResults,
                        outFileNameResult=self.outFileNameResults, outFileNameGeom=self.outFileNameLog, tt1=tt1
                    )
                    purgeFiles(endPath, partName, pathToAbaqus, jobName)
                except Exception as e:
                    change_direction()
                    try:
                        write_inp_contact(
                            fileName=inpFileName + '.inp', Nodes=shellNode, Elements=shellEle,
                            BCfix=fixed_bc, THC=THK, Emod=EM, Dens=Dens, JobName=jobName,
                            ModelName=modelName, partName=partName, MaterialName=MaterialName, PressType=PressType,
                            press_overclosure='linear', tangent_behavior=tangent_behavior,
                            normal_behavior=normal_behavior
                        )
                        message = run_abaqus(pathToAbaqus, jobName, inpFileName, self.cpus)
                        outFEATime = datetime.datetime.now() - tt1
                    except Exception as e:
                        delta = datetime.datetime.now() - tt1
                        # wbGeom = load_workbook(filename=self.outFileNameLog)
                        sheet = self.wbLog['log']
                        sheet.append(
                            [fileName, get_id(), HGT, Lstr, SEC, DIA, THK, ANG, Lift, CVT, LAS, EM, tangent_behavior,
                             normal_behavior,
                             0, str(e), delta])
                        self.wbLog.save(self.outFileNameLog)
                        self.wbLog.close()
                        del delta, sheet
                        raise e
                    #change_direction()
                    # парсим odb, считываем поля, считаем максимумы и площадь открытия, пишем в outFileName
                    try: # already in try: and switched direction
                        _ = get_history_output(pathName=pathToAbaqus, odbFileName=jobName + '.odb')
                    except:
                        delta = datetime.datetime.now() - tt1
                        # wbGeom = load_workbook(filename=self.outFileNameLog)
                        sheet = self.wbLog['log']
                        with open(pathToAbaqus + '/results/' + partName[0:-5] + '/FramesCount.txt', 'r') as DataFile:
                            frames = int(DataFile.read())
                        sheet.append(
                            [fileName, get_id(), HGT, Lstr, SEC, DIA, THK, ANG, Lift, CVT, LAS, EM, tangent_behavior,
                             normal_behavior,
                             frames,
                             'Odb parse problem', delta])
                        self.wbLog.save(self.outFileNameLog)
                        self.wbLog.close()
                        # purgeFiles(pathToAbaqus + 'results/', partName, pathToAbaqus, jobName)
                        del delta, sheet
                        raise 'Odb parse problem'

                    try:
                        endPath = pathToAbaqus + 'results/'
                        LMN_op, LMN_cl, Smax, VMS, perf_index, heli = read_data(
                            pathToAbaqus=pathToAbaqus, endPath=endPath, partName=partName, Slim=3.23,
                            HGT=HGT, Lstr=Lstr, DIA=DIA, THK=THK, ANG=ANG, Lift=Lift, CVT=CVT, LAS=LAS, EM=EM, SEC=SEC,
                            tangent_behavior=tangent_behavior, normal_behavior=normal_behavior,
                            mesh_step=mesh_step, outFEATime=outFEATime, fileName=fileName,
                            sheet_short=self.sheet_short, sheet_desc=self.sheet_desc, wbResults=self.wbResults,
                            outFileNameResult=self.outFileNameResults, outFileNameGeom=self.outFileNameLog, tt1=tt1
                        )
                        purgeFiles(endPath, partName, pathToAbaqus, jobName)
                    except:
                        delta = datetime.datetime.now() - tt1
                        # wbGeom = load_workbook(filename=self.outFileNameLog)
                        sheet = self.wbLog['log']
                        sheet.append(
                            [fileName, get_id(), HGT, Lstr, SEC, DIA, THK, ANG, Lift, CVT, LAS, EM, tangent_behavior,
                             normal_behavior,
                             0, str(e), delta])
                        self.wbLog.save(self.outFileNameLog)
                        self.wbLog.close()
                        del delta, sheet
                        raise e
                    delta = datetime.datetime.now() - tt1
                    sheet = self.wbLog['log']
                    try:
                        with open(pathToAbaqus + '/results/' + partName[0:-5].upper() + '/FramesCount.txt',
                                  'r') as DataFile:
                            frames = int(DataFile.read())
                    except:
                        frames = 0
                    sheet.append(
                        [fileName, get_id(), HGT, Lstr, SEC, DIA, THK, ANG, Lift, CVT, LAS, EM, tangent_behavior,
                         normal_behavior,
                         frames,
                         str(e), delta])
                    self.wbLog.save(self.outFileNameLog)
                    self.wbLog.close()
                    purgeFiles(pathToAbaqus + 'results/', partName, pathToAbaqus, jobName)
                    del delta, sheet
                    raise e
                del fixed_bc, partName, jobName, endPath, modelName, inpFileName
                del tt1, tt2

                objectives_dict = {
                    '1 - LMN_open': 1 - LMN_op,
                    'LMN_open': LMN_op,
                    "LMN_closed": LMN_cl,
                    "Smax - Slim": Smax - get_s_lim(),
                    'HELI': heli
                }

                constraints_dict = {
                    "VMS-Smax": VMS - get_s_lim()
                }

                # cleanup_log_leaflet()
                # set_id(ID + 1)

                return {"objectives": objectives_dict, "constraints": constraints_dict}
            except Exception as exept:
                log_message(f'Exception: {exept}')

                objectives_dict = {
                    'LMN_open': 0.0,
                    '1 - LMN_open': 2.0,
                    "LMN_closed": 2.0,
                    "Smax - Slim": 5,
                    'HELI': 3
                }

                constraints_dict = {
                    "LMN_op_constr": 0.0,
                    "1 - LMN_op_constr": 2.0,
                    "LMN_cl_constr": 2.0,
                    "Smax_constr": 50,
                    "VMS-Smax": 50
                }
                # cleanup_log_leaflet()
                # set_id(ID + 1)

                return {"objectives": objectives_dict, "constraints": constraints_dict}

        def run_pymoo(self, params) -> dict:
            problem = get_problem("welded_beam")
            param_array = np.array(list(params.values()))
            result = problem.evaluate(param_array)
            objective_values = result[0]
            objectives_dict = {
                "objective1": objective_values[0],
                "objective2": objective_values[1]
            }
            constraint_values = result[1]
            constraints_dict = {
                "constraint1": constraint_values[0],
                "constraint2": constraint_values[1],
                "constraint3": constraint_values[2],
                "constraint4": constraint_values[3]
            }

            # cleanup_log_leaflet()
            # set_id(ID + 1)

            return {"objectives": objectives_dict, "constraints": constraints_dict}

        problem_name = get_problem_name().lower()
        if problem_name == 'leaflet_single':
            res = run_leaflet_single(self, params)
        elif problem_name == 'leaflet_contact':
            res = run_leaflet_contact(self, params)
        elif problem_name == 'test':
            res = run_pymoo(self, params)

        set_id(get_id() + 1)
        return res


# ====================================================================================================================
# ___________________________________________INIT PROCEDURE___________________________________________________________
# ====================================================================================================================

# абсолютный путь к папке ипутов /inps/, расположенной в папке проекта
inpDir = str(pathlib.Path(__file__).parent.resolve()) + '/inps/'


def init_procedure(param_array):
    def init_procedure_leaf_single(cpus=10, logging=True, baseName='test', mesh_step=0.5):
        # prepare folders for xlsx, inps, logs, geoms
        folders = ['inps', 'logs', 'geoms']
        for folder in folders:
            if not os.path.exists('utils/' + folder):
                os.makedirs('utils/' + folder)
        del folders
        base_folder = str(pathlib.Path(__file__).parent.resolve()) + '/logs/'
        existing_subfolders = [f for f in os.listdir(base_folder) if os.path.isdir(os.path.join(base_folder, f))]
        current_time = datetime.datetime.now()
        folder_path = base_folder

        # имена и путь xls файлов; базовое имя для инпутов и .odb
        outFileNameGeom = folder_path + '/geom_' + str(baseName) + '_' + str(now) + '.xlsx'
        outFileNameResult = folder_path + '/odb_' + str(baseName) + '_' + str(now) + '.xlsx'
        if not glob(os.path.join(folder_path, outFileNameGeom)):
            # подготовка таблиц
            colNamesRes = pd.DataFrame(
                {'fileName': [], 'HGT, mm': [], 'Lstr, mm': [], 'SEC, deg': [], 'DIA, mm': [], 'THK, mm': [],
                 'ANG, deg': [], 'Lift, mm': [], 'CVT, %': [], 'LAS, mm': [],
                 'EM, MPa': [], 'LMN_open, mm^2/mm^2': [], 'LMN_closed, mm^2/mm^2': [], 'Smax, MPa': [], 'I': [],
                 'Tangent behavior': [], 'Normal Behavior': [],
                 'Contact Area, mm^2': [], 'ORFC, mm^2$': [], 'S_geom, mm^2': [], 'REGURG_AREA, mm^2': [], 'HELI': [],
                 'VMS, MPa': [], 'S11, MPa': [], 'S22, MPa': [], 'S33, MPa': [], 'Smid, MPa': [], 'Smin, MPa': [],
                 'LE11, m/m': [], 'LE22, m/m': [], 'LE33, m/m': [], 'LEmax, m/m': [], 'LEmid, m/m': [],
                 'LEmin, m/m': [],
                 'VMS_1, MPa': [], 'VMS_2, MPa': [], 'VMS_3, MPa': [],
                 'S11_1, MPa': [], 'S11_2, MPa': [], 'S11_3, MPa': [],
                 'S22_1, MPa': [], 'S22_2, MPa': [], 'S22_3, MPa': [],
                 'S33_1, MPa': [], 'S33_2, MPa': [], 'S33_3, MPa': [],
                 'Smax_1, MPa': [], 'Smax_2, MPa': [], 'Smax_3, MPa': [],
                 'Smid_1, MPa': [], 'Smid_2, MPa': [], 'Smid_3, MPa': [],
                 'Smin_1, MPa': [], 'Smin_2, MPa': [], 'Smin_3, MPa': [],
                 'LE11_1, m/m': [], 'LE11_2, m/m': [], 'LE11_3, m/m': [],
                 'LE22_1, m/m': [], 'LE22_2, m/m': [], 'LE22_3, m/m': [],
                 'LE33_1, m/m': [], 'LE33_2, m/m': [], 'LE33_3, m/m': [],
                 'LEmax_1, m/m': [], 'LEmax_2, m/m': [], 'LEmax_3, m/m': [],
                 'LEmid_1, m/m': [], 'LEmid_2, m/m': [], 'LEmid_3, m/m': [],
                 'LEmin_1, m/m': [], 'LEmin_2, m/m': [], 'LEmin_3, m/m': [],
                 'Elements/p1': [], 'Elements/tot': [], 'FEA time, hh:mm:ss.ms': [], 'Frames': [],
                 'Total time, hh:mm:ss.ms': []})

            colNamesRes_short = pd.DataFrame(
                {'fileName': [], 'HGT, mm': [], 'Lstr, mm': [], 'SEC, deg': [], 'DIA, mm': [], 'THK, mm': [],
                 'ANG, deg': [], 'Lift, mm': [], 'CVT, %': [], 'LAS, mm': [],
                 'EM, MPa': [], 'LMN_open, mm^2/mm^2': [], 'LMN_closed, mm^2/mm^2': [], 'Smax, MPa': [], 'I': []})

            writerRes = pd.ExcelWriter(str(outFileNameResult), engine='xlsxwriter')
            colNamesRes_short.to_excel(writerRes, sheet_name='short', index=False)
            colNamesRes.to_excel(writerRes, sheet_name='descriptive', index=False)
            writerRes._save()

            colNamesGeoms = pd.DataFrame(
                {'fileName': [], 'HGT': [], 'Lstr': [], 'SEC': [], 'DIA': [], 'THK': [],
                 'ANG': [], 'Lift': [], 'CVT': [], 'LAS': [], 'EM': [],
                 'Tangent behavior': [], 'Normal Behavior': [], 'Frames': [], 'Message': [], 'Exec time': []})

            writerGeom = pd.ExcelWriter(str(outFileNameGeom), engine='xlsxwriter')
            colNamesGeoms.to_excel(writerGeom, sheet_name='log', index=False)
            writerGeom._save()

            wbResults = load_workbook(filename=outFileNameResult)
            wbGeom = load_workbook(filename=outFileNameGeom)
            sheet_short = wbResults['short']
            sheet_desc = wbResults['descriptive']

            if logging:
                configure_log_leaflet(folder_path + '/log' + '_' + str(now))
        else:
            wbResults = load_workbook(filename=outFileNameResult)
            wbGeom = load_workbook(filename=outFileNameGeom)
            sheet_short = wbResults['short']
            sheet_desc = wbResults['descriptive']
        problem = Procedure(cpus=cpus, logging=logging, baseName=baseName, mesh_step=get_mesh_step(),
                            wbResults=wbResults, wbLog=wbGeom,
                            outFileNameResults=outFileNameResult, outFileNameLog=outFileNameGeom,
                            sheet_short=sheet_short,
                            sheet_desc=sheet_desc)

        return problem

    def init_procedure_leaf_contact(cpus=10, logging=True, baseName='test'):
        # prepare folders for xlsx, inps, logs, geoms
        folders = ['inps', 'logs', 'geoms']
        for folder in folders:
            if not os.path.exists('utils/' + folder):
                os.makedirs('utils/' + folder)
        del folders
        base_folder = str(pathlib.Path(__file__).parent.resolve()) + '/logs/'
        existing_subfolders = [f for f in os.listdir(base_folder) if os.path.isdir(os.path.join(base_folder, f))]
        current_time = datetime.datetime.now()
        folder_path = base_folder

        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
            print(f"Created folder: {folder_path}")

        # имена и путь xls файлов; базовое имя для инпутов и .odb
        outFileNameGeom = folder_path + '/geom_' + str(baseName) + '_' + str(now) + '.xlsx'
        outFileNameResult = folder_path + '/odb_' + str(baseName) + '_' + str(now) + '.xlsx'
        if not glob(os.path.join(folder_path, outFileNameGeom)):
            # подготовка таблиц
            colNamesRes = pd.DataFrame(
                {'fileName': [], 'ID':[], 'HGT, mm': [], 'Lstr, mm': [], 'SEC, deg': [], 'DIA, mm': [], 'THK, mm': [],
                 'ANG, deg': [], 'Lift, mm': [], 'CVT, %': [], 'LAS, mm': [],
                 'EM, MPa': [], 'LMN_open, mm^2/mm^2': [], 'LMN_closed, mm^2/mm^2': [], 'Smax, MPa': [], 'I': [],
                 'Tangent behavior': [], 'Normal Behavior': [],
                 'Contact Area, mm^2': [], 'ORFC, mm^2$': [], 'S_geom, mm^2': [], 'REGURG_AREA, mm^2': [], 'HELI': [],
                 'VMS, MPa': [], 'S11, MPa': [], 'S22, MPa': [], 'S33, MPa': [], 'Smid, MPa': [], 'Smin, MPa': [],
                 'LE11, m/m': [], 'LE22, m/m': [], 'LE33, m/m': [], 'LEmax, m/m': [], 'LEmid, m/m': [], 'LEmin, m/m': [],
                 'VMS_1, MPa': [], 'VMS_2, MPa': [], 'VMS_3, MPa': [],
                 'S11_1, MPa': [], 'S11_2, MPa': [], 'S11_3, MPa': [],
                 'S22_1, MPa': [], 'S22_2, MPa': [], 'S22_3, MPa': [],
                 'S33_1, MPa': [], 'S33_2, MPa': [], 'S33_3, MPa': [],
                 'Smax_1, MPa': [], 'Smax_2, MPa': [], 'Smax_3, MPa': [],
                 'Smid_1, MPa': [], 'Smid_2, MPa': [], 'Smid_3, MPa': [],
                 'Smin_1, MPa': [], 'Smin_2, MPa': [], 'Smin_3, MPa': [],
                 'LE11_1, m/m': [], 'LE11_2, m/m': [], 'LE11_3, m/m': [],
                 'LE22_1, m/m': [], 'LE22_2, m/m': [], 'LE22_3, m/m': [],
                 'LE33_1, m/m': [], 'LE33_2, m/m': [], 'LE33_3, m/m': [],
                 'LEmax_1, m/m': [], 'LEmax_2, m/m': [], 'LEmax_3, m/m': [],
                 'LEmid_1, m/m': [], 'LEmid_2, m/m': [], 'LEmid_3, m/m': [],
                 'LEmin_1, m/m': [], 'LEmin_2, m/m': [], 'LEmin_3, m/m': [],
                 'Elements/p1': [], 'Elements/tot': [], 'FEA time, hh:mm:ss.ms': [], 'Frames': [],
                 'Total time, hh:mm:ss.ms': []})

            colNamesRes_short = pd.DataFrame(
                {'fileName': [], 'ID':[], 'HGT, mm': [], 'Lstr, mm': [], 'SEC, deg': [], 'DIA, mm': [], 'THK, mm': [],
                 'ANG, deg': [], 'Lift, mm': [], 'CVT, %': [], 'LAS, mm': [],
                 'EM, MPa': [], 'LMN_open, mm^2/mm^2': [], 'LMN_closed, mm^2/mm^2': [], 'Smax, MPa': [], 'I': []})

            writerRes = pd.ExcelWriter(str(outFileNameResult), engine='xlsxwriter')
            colNamesRes_short.to_excel(writerRes, sheet_name='short', index=False)
            colNamesRes.to_excel(writerRes, sheet_name='descriptive', index=False)
            writerRes._save()

            colNamesGeoms = pd.DataFrame(
                {'fileName': [], 'ID':[], 'HGT': [], 'Lstr': [], 'SEC': [], 'DIA': [], 'THK': [],
                 'ANG': [], 'Lift': [], 'CVT': [], 'LAS': [], 'EM': [],
                 'Tangent behavior': [], 'Normal Behavior': [], 'Frames': [], 'Message': [], 'Exec time': []})

            writerGeom = pd.ExcelWriter(str(outFileNameGeom), engine='xlsxwriter')
            colNamesGeoms.to_excel(writerGeom, sheet_name='log', index=False)
            writerGeom._save()

            wbResults = load_workbook(filename=outFileNameResult)
            wbGeom = load_workbook(filename=outFileNameGeom)
            sheet_short = wbResults['short']
            sheet_desc = wbResults['descriptive']

            if logging:
                configure_log_leaflet(folder_path + '/log' + '_' + str(now))

        else:
            wbResults = load_workbook(filename=outFileNameResult)
            wbGeom = load_workbook(filename=outFileNameGeom)
            sheet_short = wbResults['short']
            sheet_desc = wbResults['descriptive']

        problem = Procedure(cpus=cpus, logging=logging, baseName=baseName, mesh_step=get_mesh_step(),
                            wbResults=wbResults,
                            wbLog=wbGeom,
                            outFileNameResults=outFileNameResult, outFileNameLog=outFileNameGeom,
                            sheet_short=sheet_short,
                            sheet_desc=sheet_desc)

        return problem

    problem_name = get_problem_name().lower()
    cpus = get_cpus()
    parameters = param_array
    problem = None
    if problem_name == 'leaflet_single':
        problem = init_procedure_leaf_single(cpus=cpus, logging=True, baseName=get_base_name())
    elif problem_name == 'leaflet_contact':
        problem = init_procedure_leaf_contact(cpus=cpus, logging=True, baseName=get_base_name())
    elif problem_name == 'test':
        problem = get_problem("welded_beam")
    else:
        raise Exception(f'Wrong problem name: {problem_name}, you entered: {problem_name}')
    return problem
