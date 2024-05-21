import numpy as np
from glob2 import glob
import os
import datetime
from openpyxl import load_workbook
from .computeGap import *
from .helicopter import helicopter
from .global_variable import get_problem_name
from .logger_leaflet import log_message


# ====================================================================================================================
# _____________________________________________CLASS__________________________________________________________________
# ====================================================================================================================

# обертка для хранения и обработки результатов парсинга odb
# для сетки
class readedMesh:
    def __init__(self, lnodes, lelements):
        self.nodes = lnodes
        self.elements = lelements

    def repNodes(self, lnodes):
        self.nodes = lnodes

    def repElements(self, lelements):
        self.elements = lelements

    def addNodes(self, lnodes):
        self.nodes = np.append(self.nodes, lnodes)

    def addElements(self, lelements):
        self.elements = np.append(self.elements, lelements)

    def getNodes(self):
        return self.nodes

    def getElements(self):
        return self.elements


def read_data_beam(pathToAbaqus=None, endPath=None, partName=None, ID=-1,
              Width=-1, THK=-1, fileName='', frames=-1, delta=-1,
              sheet_short=None, sheet_desc=None, wbResults=None, outFileNameResult=None,
              outFileNameGeom=None):
    def read(pathToAbaqus=None, endPath=None, partName=None):
        foldPath = endPath + partName[0:-5].upper() + '/' + partName.upper()

        mesh1 = readedMesh(lnodes=np.loadtxt(foldPath + '-1/Nodes.txt'),
                           lelements=np.loadtxt(foldPath + '-1/Elements.txt'))
        stressFiles = glob(foldPath + '-1/' + 'Stress_*.txt')

        newstr = ''.join((ch if ch in '0123456789' else ' ') for ch in stressFiles[0])
        one_step = ([int(i) for i in newstr.split()])[-1]

        s_max1_1step = np.loadtxt(foldPath + '-1/' + 'Stress_' + str(one_step) + '.txt')

        eleCount = len(mesh1.getElements())

        disp1 = np.loadtxt(foldPath + '-1/' + 'U_' + str(one_step) + '.txt')

        return (s_max1_1step, disp1)

    subfolders = [f.path for f in os.scandir(endPath + partName[:-5].upper()) if f.is_dir()]

    (s_max1_1step, disp1) = read(
        pathToAbaqus=pathToAbaqus, endPath=endPath, partName=partName
    )
    VMS = s_max1_1step[-1]
    S11 = s_max1_1step[0]
    S22 = s_max1_1step[1]
    Smax = s_max1_1step[3]
    Smid = s_max1_1step[4]
    Smin = s_max1_1step[5]

    disp1 = disp1[-1][2]

    new_row_short = [fileName, ID, Width, THK, disp1]
    new_row_desc = [
        fileName, ID, Width, THK, disp1, Smax,
        VMS, S11, S22, Smid, Smin,
        s_max1_1step[-1],  # VMS
        s_max1_1step[0],  # S11
        s_max1_1step[1],  # S22
        s_max1_1step[3],  # Smax
        s_max1_1step[4],  # Smid
        s_max1_1step[5],  # Smin
        frames, delta
    ]

    sheet_short.append(new_row_short)
    sheet_desc.append(new_row_desc)
    wbResults.save(outFileNameResult)
    wbResults.close()

    wbGeom = load_workbook(filename=outFileNameGeom)
    sheet = wbGeom['log']
    sheet.append([fileName, ID, Width, THK, frames, str('Done'), delta])
    wbGeom.save(outFileNameGeom)
    wbGeom.close()

    return Smax, disp1




def read_data_leaf(pathToAbaqus=None, endPath=None, partName=None, Slim=9.9,
              HGT=-1, Lstr=-1, DIA=10, THK=-1, ANG=-1, Lift=-1, CVT=-1, LAS=-1, EM=-1,  SEC=120,
              tangent_behavior=0, normal_behavior=0, mesh_step=0.35, outFEATime=-1, fileName='',
              sheet_short=None, sheet_desc=None, wbResults=None, outFileNameResult=None,
              outFileNameGeom=None, tt1=None):
    def read_data_contact(pathToAbaqus=None, endPath=None, partName=None, DIA=10, SEC=120, mesh_step=0.35):
        foldPath = endPath + partName[0:-5].upper() + '/' + partName.upper()

        mesh1 = readedMesh(lnodes=np.loadtxt(foldPath + '-1/Nodes.txt'),
                           lelements=np.loadtxt(foldPath + '-1/Elements.txt'))
        mesh2 = readedMesh(lnodes=np.loadtxt(foldPath + '-2/Nodes.txt'),
                           lelements=np.loadtxt(foldPath + '-2/Elements.txt'))
        mesh3 = readedMesh(lnodes=np.loadtxt(foldPath + '-3/Nodes.txt'),
                           lelements=np.loadtxt(foldPath + '-3/Elements.txt'))
        stressFiles = glob(foldPath + '-1/' + 'Stress_*.txt')

        newstr = ''.join((ch if ch in '0123456789' else ' ') for ch in stressFiles[0])
        one_step = ([int(i) for i in newstr.split()])[-1]

        newstr = ''.join((ch if ch in '0123456789' else ' ') for ch in stressFiles[1])
        two_step = ([int(i) for i in newstr.split()])[-1]

        closed_step = min(one_step, two_step)
        opened_step = max(one_step, two_step)

        del newstr, one_step, two_step

        s_max1_1step = np.loadtxt(foldPath + '-1/' + 'Stress_' + str(closed_step) + '.txt')
        s_max1_2step = np.loadtxt(foldPath + '-1/' + 'Stress_' + str(opened_step) + '.txt')
        le_max1_1step = np.loadtxt(foldPath + '-1/' + 'LE_' + str(closed_step) + '.txt')
        le_max1_2step = np.loadtxt(foldPath + '-1/' + 'LE_' + str(opened_step) + '.txt')

        s_max2_1step = np.loadtxt(foldPath + '-2/' + 'Stress_' + str(closed_step) + '.txt')
        s_max2_2step = np.loadtxt(foldPath + '-2/' + 'Stress_' + str(opened_step) + '.txt')
        le_max2_1step = np.loadtxt(foldPath + '-2/' + 'LE_' + str(closed_step) + '.txt')
        le_max2_2step = np.loadtxt(foldPath + '-2/' + 'LE_' + str(opened_step) + '.txt')

        s_max3_1step = np.loadtxt(foldPath + '-3/' + 'Stress_' + str(closed_step) + '.txt')
        s_max3_2step = np.loadtxt(foldPath + '-3/' + 'Stress_' + str(opened_step) + '.txt')
        le_max3_1step = np.loadtxt(foldPath + '-3/' + 'LE_' + str(closed_step) + '.txt')
        le_max3_2step = np.loadtxt(foldPath + '-3/' + 'LE_' + str(opened_step) + '.txt')

        eleCount = len(mesh1.getElements())

        disp1 = np.loadtxt(foldPath + '-1/' + 'U_' + str(opened_step) + '.txt')
        disp2 = np.loadtxt(foldPath + '-2/' + 'U_' + str(opened_step) + '.txt')
        disp3 = np.loadtxt(foldPath + '-3/' + 'U_' + str(opened_step) + '.txt')
        try:
            outGap1 = computeOpened(disp1[:, 1:], mesh1.nodes[:, 1:], mesh_step=mesh_step)
        except:
            disp1 = disp1[:len(mesh1.nodes[:, 1:]), :]
            outGap1 = computeOpened(disp1[:, 1:], mesh1.nodes[:, 1:], mesh_step=mesh_step)
            disp2 = disp2[:len(mesh2.nodes[:, 1:]), :]
            disp3 = disp3[:len(mesh3.nodes[:, 1:]), :]

        outGap2 = computeOpened(disp2[:len(mesh2.nodes[:, 1:]), 1:], mesh2.nodes[:, 1:], mesh_step=mesh_step)
        outGap3 = computeOpened(disp3[:len(mesh3.nodes[:, 1:]), 1:], mesh3.nodes[:, 1:], mesh_step=mesh_step)

        maxORFC = np.pi * (DIA / 2) * (DIA / 2)
        opened_area = maxORFC - (outGap1 + outGap2 + outGap3)
        LMN_op = opened_area / maxORFC
        disp1 = np.loadtxt(foldPath + '-1/' + 'U_' + str(closed_step) + '.txt')
        disp2 = np.loadtxt(foldPath + '-2/' + 'U_' + str(closed_step) + '.txt')
        disp3 = np.loadtxt(foldPath + '-3/' + 'U_' + str(closed_step) + '.txt')
        closed_area = computeClosed(disp1=disp1[:, 1:], disp2=disp2[:, 1:], disp3=disp3[:, 1:],
                                    nodes1=mesh1.nodes[:, 1:],
                                    nodes2=mesh2.nodes[:, 1:], nodes3=mesh3.nodes[:, 1:], mesh_step=mesh_step)
        LMN_cl = (maxORFC - closed_area) / maxORFC
        cos120pl, cos120min = np.cos(np.deg2rad(120)), np.cos(np.deg2rad(-120))
        sin120pl, sin120min = np.sin(np.deg2rad(120)), np.sin(np.deg2rad(-120))
        rotMatrixPlus = [[cos120pl, -sin120pl, 0], [sin120pl, cos120pl, 0], [0, 0, 1]]
        rotMatrixMinus = [[cos120min, -sin120min, 0], [sin120min, cos120min, 0], [0, 0, 1]]
        isHelicopter = helicopter(
            nodes=(mesh1.nodes[:, 1:] + disp1[:len(mesh1.nodes[:, 1:]), 1:]), SEC=SEC
        ) + helicopter(
            nodes=np.dot(mesh2.nodes[:, 1:] + disp2[:len(mesh2.nodes[:, 1:]), 1:], rotMatrixMinus), SEC=SEC
        ) + helicopter(
            nodes=np.dot(mesh3.nodes[:, 1:] + disp3[:len(mesh3.nodes[:, 1:]), 1:], rotMatrixPlus), SEC=SEC
        )
        del outGap1, outGap2, outGap3, mesh1, mesh2, mesh3, disp1, disp2, disp3, closed_step, opened_step
        del cos120pl, cos120min, sin120min, sin120pl, rotMatrixMinus, rotMatrixPlus
        try:
            carea = 0
            for file in glob(endPath + partName[0:-5].upper() + '/carea_*.txt'):
                t_carea = np.loadtxt(file)
                carea += max(t_carea.T[1])
            del t_carea
        except Exception as ex:
            raise 'Uncomplited odb. Frames < 21'
        try:
            with open(pathToAbaqus + '/results/' + partName[0:-5].upper() + '/FramesCount.txt', 'r') as DataFile:
                frames = int(DataFile.read())
        except:
            frames = 0

        return (s_max1_1step, s_max1_2step, le_max1_1step, le_max1_2step, s_max2_1step, s_max2_2step, le_max2_1step,
                le_max2_2step, s_max3_1step, s_max3_2step, le_max3_1step, le_max3_2step, eleCount, LMN_op, LMN_cl,
                isHelicopter, frames, opened_area, closed_area, maxORFC, carea)

    def read_data_single(pathToAbaqus=None, endPath=None, partName=None, DIA=10, SEC=120, mesh_step=0.35):
        foldPath = endPath + partName[0:-5].upper() + '/' + partName.upper()

        mesh1 = readedMesh(lnodes=np.loadtxt(foldPath + '-1/Nodes.txt'),
                           lelements=np.loadtxt(foldPath + '-1/Elements.txt'))
        stressFiles = glob(foldPath + '-1/' + 'Stress_*.txt')

        newstr = ''.join((ch if ch in '0123456789' else ' ') for ch in stressFiles[0])
        one_step = ([int(i) for i in newstr.split()])[-1]

        newstr = ''.join((ch if ch in '0123456789' else ' ') for ch in stressFiles[1])
        two_step = ([int(i) for i in newstr.split()])[-1]

        closed_step = min(one_step, two_step)
        opened_step = max(one_step, two_step)

        del newstr, one_step, two_step

        s_max1_1step = np.loadtxt(foldPath + '-1/' + 'Stress_' + str(closed_step) + '.txt')
        s_max1_2step = np.loadtxt(foldPath + '-1/' + 'Stress_' + str(opened_step) + '.txt')
        le_max1_1step = np.loadtxt(foldPath + '-1/' + 'LE_' + str(closed_step) + '.txt')
        le_max1_2step = np.loadtxt(foldPath + '-1/' + 'LE_' + str(opened_step) + '.txt')

        eleCount = len(mesh1.getElements())

        disp1 = np.loadtxt(foldPath + '-1/' + 'U_' + str(opened_step) + '.txt')

        try:
            outGap1 = computeOpened(disp1[:, 1:], mesh1.nodes[:, 1:], mesh_step=mesh_step)
        except:
            disp1 = disp1[:len(mesh1.nodes[:, 1:]), :]
            outGap1 = computeOpened(disp1[:, 1:], mesh1.nodes[:, 1:], mesh_step=mesh_step)

        maxORFC = np.pi * (DIA / 2) * (DIA / 2)
        opened_area = maxORFC - (3 * outGap1)
        LMN_op = opened_area / maxORFC

        closed_area = computeClosed_single(disp1=disp1[:, 1:], nodes1=mesh1.nodes[:, 1:], mesh_step=mesh_step)
        LMN_cl = (maxORFC - closed_area) / maxORFC
        del outGap1, mesh1, disp1, closed_step, opened_step

        try:
            with open(pathToAbaqus + '/results/' + partName[0:-5].upper() + '/FramesCount.txt', 'r') as DataFile:
                frames = int(DataFile.read())
        except:
            frames = 0

        return (s_max1_1step, s_max1_2step, le_max1_1step, le_max1_2step, eleCount, LMN_op, LMN_cl, frames, opened_area,
                closed_area, maxORFC)

    subfolders = [f.path for f in os.scandir(endPath + partName[:-5].upper()) if f.is_dir()]
    if len(subfolders) > 1:
        (s_max1_1step, s_max1_2step, le_max1_1step, le_max1_2step, s_max2_1step, s_max2_2step, le_max2_1step,
         le_max2_2step, s_max3_1step, s_max3_2step, le_max3_1step, le_max3_2step, eleCount, LMN_op, LMN_cl,
         isHelicopter, frames, opened_area, closed_area, maxORFC, carea) = read_data_contact(
            pathToAbaqus=pathToAbaqus, endPath=endPath, partName=partName, DIA=DIA, SEC=SEC, mesh_step=mesh_step
        )
        VMS = (max(s_max1_1step[-1], s_max1_2step[-1]) + max(s_max2_1step[-1], s_max2_2step[-1])
               + max(s_max3_1step[-1], s_max3_2step[-1])) / 3
        S11 = (max(s_max1_1step[0], s_max1_2step[0]) + max(s_max2_1step[0], s_max2_2step[0])
               + max(s_max3_1step[0], s_max3_2step[0])) / 3
        S22 = (max(s_max1_1step[1], s_max1_2step[1]) + max(s_max2_1step[1], s_max2_2step[1])
               + max(s_max3_1step[1], s_max3_2step[1])) / 3
        S33 = (max(s_max1_1step[2], s_max1_2step[2]) + max(s_max2_1step[2], s_max2_2step[2])
               + max(s_max3_1step[2], s_max3_2step[2])) / 3
        Smax = (max(s_max1_1step[3], s_max1_2step[3]) + max(s_max2_1step[3], s_max2_2step[3])
                + max(s_max3_1step[3], s_max3_2step[3])) / 3
        Smid = (max(s_max1_1step[4], s_max1_2step[4]) + max(s_max2_1step[4], s_max2_2step[4])
                + max(s_max3_1step[4], s_max3_2step[4])) / 3
        Smin = (max(s_max1_1step[5], s_max1_2step[5]) + max(s_max2_1step[5], s_max2_2step[5])
                + max(s_max3_1step[5], s_max3_2step[5])) / 3

        LE11 = (max(le_max1_1step[0], le_max1_2step[0]) + max(le_max2_1step[0], le_max2_2step[0])
                + max(le_max3_1step[0], le_max3_2step[0])) / 3
        LE22 = (max(le_max1_1step[1], le_max1_2step[1]) + max(le_max2_1step[1], le_max2_2step[1])
                + max(le_max3_1step[1], le_max3_2step[1])) / 3
        LE33 = (max(le_max1_1step[2], le_max1_2step[2]) + max(le_max2_1step[2], le_max2_2step[2])
                + max(le_max3_1step[2], le_max3_2step[2])) / 3
        LEmax = (max(le_max1_1step[3], le_max1_2step[3]) + max(le_max2_1step[3], le_max2_2step[3])
                 + max(le_max3_1step[3], le_max3_2step[3])) / 3
        LEmid = (max(le_max1_1step[4], le_max1_2step[4]) + max(le_max2_1step[4], le_max2_2step[4])
                 + max(le_max3_1step[4], le_max3_2step[4])) / 3
        LEmin = (max(le_max1_1step[5], le_max1_2step[5]) + max(le_max2_1step[5], le_max2_2step[5])
                 + max(le_max3_1step[5], le_max3_2step[5])) / 3
    else:
        (s_max1_1step, s_max1_2step, le_max1_1step, le_max1_2step, eleCount, LMN_op, LMN_cl, frames, opened_area,
         closed_area, maxORFC) = read_data_single(
            pathToAbaqus=pathToAbaqus, endPath=endPath, partName=partName, DIA=DIA, SEC=SEC, mesh_step=mesh_step
        )
        VMS = max(s_max1_1step[-1], s_max1_2step[-1])
        S11 = max(s_max1_1step[0], s_max1_2step[0])
        S22 = max(s_max1_1step[1], s_max1_2step[1])
        S33 = max(s_max1_1step[2], s_max1_2step[2])
        Smax = max(s_max1_1step[3], s_max1_2step[3])
        Smid = max(s_max1_1step[4], s_max1_2step[4])
        Smin = max(s_max1_1step[5], s_max1_2step[5])

        LE11 = max(le_max1_1step[0], le_max1_2step[0])
        LE22 = max(le_max1_1step[1], le_max1_2step[1])
        LE33 = max(le_max1_1step[2], le_max1_2step[2])
        LEmax = max(le_max1_1step[3], le_max1_2step[3])
        LEmid = max(le_max1_1step[4], le_max1_2step[4])
        LEmin = max(le_max1_1step[5], le_max1_2step[5])

    tdiff1 = datetime.datetime.now() - tt1

    if Smax < Slim:
        perf_index = np.round(np.sqrt(
            np.power((opened_area / maxORFC), 2)
            + np.power(1 - ((maxORFC - closed_area) / maxORFC), 2)
            + np.power((1 - (Smax / 9.9)), 2)
        ), 4)
    else:
        perf_index = 0

    if len(subfolders)> 1:
        new_row_short = [fileName, HGT, Lstr, SEC, DIA, THK, ANG, Lift, CVT, LAS, EM, np.round(LMN_op, 3),
                         np.round(LMN_cl, 3), np.round(Smax, 4), perf_index]
        new_row_desc = [fileName, HGT, Lstr, SEC, DIA, THK, ANG, Lift, CVT, LAS, EM, LMN_op, LMN_cl, Smax,
                        perf_index,
                        tangent_behavior, normal_behavior, carea, opened_area, maxORFC, (maxORFC - closed_area),
                        isHelicopter,
                        VMS, S11, S22, S33, Smid, Smin, LE11, LE22, LE33, LEmax, LEmid, LEmin,
                        max(s_max1_1step[-1], s_max1_2step[-1]), max(s_max2_1step[-1], s_max2_2step[-1]),
                        max(s_max3_1step[-1], s_max3_2step[-1]),  # VMS
                        max(s_max1_1step[0], s_max1_2step[0]), max(s_max2_1step[0], s_max2_2step[0]),
                        max(s_max3_1step[0], s_max3_2step[0]),  # S11
                        max(s_max1_1step[1], s_max1_2step[1]), max(s_max2_1step[1], s_max2_2step[1]),
                        max(s_max3_1step[1], s_max3_2step[1]),  # S22
                        max(s_max1_1step[2], s_max1_2step[2]), max(s_max2_1step[2], s_max2_2step[2]),
                        max(s_max3_1step[2], s_max3_2step[2]),  # S33
                        max(s_max1_1step[3], s_max1_2step[3]), max(s_max2_1step[3], s_max2_2step[3]),
                        max(s_max3_1step[3], s_max3_2step[3]),  # Smax
                        max(s_max1_1step[4], s_max1_2step[4]), max(s_max2_1step[4], s_max2_2step[4]),
                        max(s_max3_1step[4], s_max3_2step[4]),  # Smid
                        max(s_max1_1step[5], s_max1_2step[5]), max(s_max2_1step[5], s_max2_2step[5]),
                        max(s_max3_1step[5], s_max3_2step[5]),  # Smin
                        max(le_max1_1step[0], le_max1_2step[0]), max(le_max2_1step[0], le_max2_2step[0]),
                        max(le_max3_1step[0], le_max3_2step[0]),  # LE11
                        max(le_max1_1step[1], le_max1_2step[1]), max(le_max2_1step[1], le_max2_2step[1]),
                        max(le_max3_1step[1], le_max3_2step[1]),  # LE22
                        max(le_max1_1step[2], le_max1_2step[2]), max(le_max2_1step[2], le_max2_2step[2]),
                        max(le_max3_1step[2], le_max3_2step[2]),  # LE33
                        max(le_max1_1step[3], le_max1_2step[3]), max(le_max2_1step[3], le_max2_2step[3]),
                        max(le_max3_1step[3], le_max3_2step[3]),  # LEmax
                        max(le_max1_1step[4], le_max1_2step[4]), max(le_max2_1step[4], le_max2_2step[4]),
                        max(le_max3_1step[4], le_max3_2step[4]),  # LEmid
                        max(le_max1_1step[5], le_max1_2step[5]), max(le_max2_1step[5], le_max2_2step[5]),
                        max(le_max3_1step[5], le_max3_2step[5]),  # LEmin
                        eleCount, 3 * eleCount, outFEATime, frames, tdiff1]
    else:
        new_row_short = [fileName, HGT, Lstr, SEC, DIA, THK, ANG, Lift, CVT, LAS, EM, np.round(LMN_op, 3),
                         np.round(LMN_cl, 3), np.round(Smax, 4), perf_index]
        new_row_desc = [fileName, HGT, Lstr, SEC, DIA, THK, ANG, Lift, CVT, LAS, EM, LMN_op, LMN_cl, Smax,
                        perf_index, tangent_behavior, normal_behavior, -1, opened_area, maxORFC,
                        (maxORFC - closed_area), -1,
                        VMS, S11, S22, S33, Smid, Smin, LE11, LE22, LE33, LEmax, LEmid, LEmin,
                        max(s_max1_1step[-1], s_max1_2step[-1]), -1, -1,  # VMS
                        max(s_max1_1step[0], s_max1_2step[0]), -1, -1,  # S11
                        max(s_max1_1step[1], s_max1_2step[1]), -1, -1,  # S22
                        max(s_max1_1step[2], s_max1_2step[2]), -1, -1,  # S33
                        max(s_max1_1step[3], s_max1_2step[3]), -1, - 1,  # Smax
                        max(s_max1_1step[4], s_max1_2step[4]), -1, -1,  # Smid
                        max(s_max1_1step[5], s_max1_2step[5]), -1, -1,  # Smin
                        max(le_max1_1step[0], le_max1_2step[0]), -1, -1,  # LE11
                        max(le_max1_1step[1], le_max1_2step[1]), -1, -1,  # LE22
                        max(le_max1_1step[2], le_max1_2step[2]), -1, -1,  # LE33
                        max(le_max1_1step[3], le_max1_2step[3]), -1, -1,  # LEmax
                        max(le_max1_1step[4], le_max1_2step[4]), -1, -1,  # LEmid
                        max(le_max1_1step[5], le_max1_2step[5]), -1, -1,  # LEmin
                        eleCount, 3 * eleCount, outFEATime, frames, tdiff1]

    sheet_short.append(new_row_short)
    sheet_desc.append(new_row_desc)
    wbResults.save(outFileNameResult)
    wbResults.close()

    wbGeom = load_workbook(filename=outFileNameGeom)
    sheet = wbGeom['log']
    sheet.append([fileName, HGT, Lstr, SEC, DIA, THK, ANG, Lift, CVT, LAS, EM,
                  tangent_behavior, normal_behavior, frames, str('Done'), tdiff1])
    wbGeom.save(outFileNameGeom)
    wbGeom.close()

    return LMN_op, LMN_cl, Smax, perf_index

def read_data(pathToAbaqus=None, endPath=None, partName=None, Width=-1, ID=-1, frames=-1, Slim=9.9,
              HGT=-1, Lstr=-1, DIA=10, THK=-1, ANG=-1, Lift=-1, CVT=-1, LAS=-1, EM=-1,  SEC=120,
              tangent_behavior=0, normal_behavior=0, mesh_step=0.35, outFEATime=-1, fileName='',
              sheet_short=None, sheet_desc=None, wbResults=None, outFileNameResult=None,
              outFileNameGeom=None, tt1=None):
    problem_name = get_problem_name().lower()

    if problem_name == 'beam':
        return read_data_beam(
            pathToAbaqus=pathToAbaqus, endPath=endPath, partName=partName, ID=ID,
            Width=Width, THK=THK, fileName=fileName, frames=frames, delta=(datetime.datetime.now() - tt1),
            sheet_short=sheet_short, sheet_desc=sheet_desc, wbResults=wbResults, outFileNameResult=outFileNameResult,
            outFileNameGeom=outFileNameGeom
        )
    elif problem_name == 'leaflet_single' or problem_name == 'leaflet_contact':
        return read_data_leaf(
            pathToAbaqus=pathToAbaqus, endPath=endPath, partName=partName, Slim=Slim,
            HGT=HGT, Lstr=Lstr, DIA=DIA, THK=THK, ANG=ANG, Lift=Lift, CVT=CVT, LAS=LAS, EM=EM,  SEC=SEC,
            tangent_behavior=tangent_behavior, normal_behavior=normal_behavior, mesh_step=mesh_step,
            outFEATime=outFEATime, fileName=fileName, sheet_short=sheet_short, sheet_desc=sheet_desc,
            wbResults=wbResults, outFileNameResult=outFileNameResult, outFileNameGeom=outFileNameGeom, tt1=tt1
        )